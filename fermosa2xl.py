from bs4 import BeautifulSoup
from openpyxl import Workbook
from threading import Thread
import requests
import re  
import time 
import sys
import datetime
import os


class Sansevierias:
    def __init__(self, base_url, ws=False):
        '''
        Class object thats used to scrape the fermosa webpage\n
        Setup base url for scraping\n
        An opened worksheet for appending scraped data\n
        Instance attributes: max_names, base_url, domain, patterns
        '''
        self.max_names = 0
        self.domain = "https://fermosaplants.com"
        self.base_url = base_url
        self.ws = ws
        self.patterns = {
            #get if a plant is variegated
            'is_Variegated': re.compile(r'variegated', re.IGNORECASE).search, 

            #get all the listing types from given types
            'listing_type': re.compile(r'combo|clump|Leaf|plant|pub', re.IGNORECASE).findall, 

            #get a floating point number from entered price
            'extract_price': re.compile(r'(Rs\.|\s|,)').sub, 

            #try to fetch name for a single listing
            'extract_name': re.compile(r'''(?:About Sansevieria (?P<name>[\-'"a-zA-Z ]*)(?: )-)''', re.IGNORECASE).search, 

            #second method to fetch name for a single listing
            'extract_namee': re.compile(r'''(?:Sansevieria (?P<namee>\w+))''').search, 

            #extract nmes from a combo offer listing
            'extract_names': re.compile(r'\d+\. *(?P<name>[\w]+(?!\.)(?: *(?!\d)[\w])*)', re.IGNORECASE).findall 
        }
        self.name_set = set()


    def fetch_from_page_url(self, url):
        '''This function calls the _get_soup and scrape page from a single method call'''
        soup = self._get_soup(url)
        page = self.scrape_page(soup)
        print(f'{len(page)} entries fetched from page {url[-1]}')


    def _get_soup(self, page_url):
        res = requests.get(page_url)
        if res.status_code==200:
            soup = BeautifulSoup(res.content, 'html.parser')
            return soup
        else:
            raise ValueError(f'Got a bad response: {res.status_code}')



    def scrape_page(self, soup:BeautifulSoup):
        '''
        Scrape page function, takes in a fermosa webpage url and scrapes listed products\n
        returns a list of scraped products data
        '''

        sansevierias = soup.find_all("div", class_="product-item-v5")
        
        #raise error if no items are fetched
        assert len(sansevierias)>0
        data_scraped = list()

        #simple function that returns a true value or the object itself if object exists otherwise a false value
        get_or_default = lambda obj, true, false: true or obj if obj else false

        for lising in sansevierias:

            _listing_name = lising.find('h4', class_='title-product')
            listing_name = _listing_name.text.strip()

            _listing_price = lising.find('span', class_='price')
            listing_price = self.patterns['extract_price']('', _listing_price.text.strip())

            listing_url_ref = _listing_name.find('a')['href'] #only 1 level below _listing_name object so use it as the parent
            listing_url = self.domain + listing_url_ref

            is_Variegated = get_or_default(self.patterns['is_Variegated'](listing_name), True, False) 
            listing_types = get_or_default(self.patterns['listing_type'](listing_name), None, [])

            item_soup = self._get_soup(listing_url)

            names = self.extract_names(item_soup, 'combo' in listing_name.lower())
            if names == []:
                names.append(listing_name.strip())
            self.name_set = self.name_set.union([name.replace("'", '').title() for name in names])

            self.max_names = max(self.max_names, len(names))

            listing_data = [
                listing_name, 
                listing_price, 
                is_Variegated,
                len(names), 
                ', '.join(set(types.lower() for types in listing_types)), 
                listing_url
                ]

            listing_data.extend([name.replace("'", '').title() for name in names])  

            if self.ws:          
                self.ws.append(listing_data)
            data_scraped.append(listing_data)

        return data_scraped
    

    def extract_names(self, soup:BeautifulSoup, combo:bool)->list[str]:
        '''
        This function extracts names from a given fermosa product webpage url,\n
        Also Information weather its a single product or a combo and returns a list of names for the listing
        '''
        if not combo:
            _summary = soup.find('div', class_='pd_summary').text
            summary = re.sub(' ', ' ', _summary)
            try:
                name = self.patterns['extract_name'](summary)
                if name:
                    return [name.group('name')]
                else:
                    return [self.patterns['extract_namee'](summary).group('namee')]
            except AttributeError:
                return []
        
        else:
            _description = soup.find('div', class_='desc product-desc').text
            description = re.sub(' ', ' ', _description)
            names = self.patterns['extract_names'](description)
            return sorted(names)


    def scrape_from(self, page_number):
        '''Simple function, takes in a page number and calls scrape_page function until products are exhausted'''
        while True:
            try:
                curr_url = self.base_url+str(page_number)
                soup = self._get_soup(curr_url)
                data_entries = self.scrape_page(soup)
                print(f'{len(data_entries)} entries fetched from page {page_number}')
                page_number+=1
            except AssertionError:
                #No more products left for fetching, add headers and return
                try:
                    headers = ['Product Name', 'Price', 'Variegated', 'Combo Amount', 'Listing Tags', 'Product Url']
                    self.add_headers(headers)
                except ValueError as e:
                    print(e)
                break
        return None


    def add_headers(self, headers:list):
        '''This function adds an empty row at the top and adds valid headers to it'''
        if not self.ws: 
            raise ValueError(f'Instance doesnt have a worksheet')

        headers.extend([f'name{n}' for n in range(1, self.max_names+1)])
        self.ws.insert_rows(1)
        for col, heading in enumerate(headers):
            self.ws.cell(row=1, column=col+1, value=heading)
        return None


def resolve_path(name, dir_path):
    return f'{dir_path}/{datetime.date.today()}-{name}.xlsx'


def main_without_threading(dir_path, name):
    wb = Workbook()
    ws = wb.create_sheet("Non Threading output")

    base_url = "https://fermosaplants.com/collections/sansevieria?page="

    fermosa_scraper = Sansevierias(base_url, ws)

    fermosa_scraper.scrape_from(page_number=1)
    print(f'{len(fermosa_scraper.name_set)} distinct values found!\n')
    # for name in fermosa_scraper.name_set:
    #     print(name, end=', ')

    wb.save(f'{dir_path}/{datetime.date.today()}-{name}.xlsx')
    


def main_with_threading(dir_path, name):
    wb = Workbook()
    ws = wb.create_sheet("Sansevierias")

    base_url = "https://fermosaplants.com/collections/sansevieria?page="
    fermosa_scraper = Sansevierias(base_url, ws)
    threads = []


    for page in range(1,8):
        curr_thread = Thread(target=fermosa_scraper.fetch_from_page_url, args=(base_url+str(page),))
        threads.append(curr_thread)
        curr_thread.start()
    
    for thread in threads:
        thread.join()

    headers = ['Product Name', 'Price', 'Variegated', 'Combo Amount', 'Listing Tags', 'Product Url']
    
    fermosa_scraper.add_headers(headers)
        
    print(f'{len(fermosa_scraper.name_set)} distinct values found!')
    wb.save(f'{dir_path}/{datetime.date.today()}-{name}.xlsx')  


if __name__ == '__main__':
    use_threading = '--no-threading' not in sys.argv
    result_dir = 'results'
    os.makedirs(result_dir, exist_ok=True)
    
    if use_threading:
        start = time.time()
        main_with_threading(result_dir, 'plantbook')
        end = time.time()
    else:
        start = time.time()
        main_without_threading(result_dir, 'plantbook')
        end = time.time()
    print(f'Script took {end-start:0.2f} seconds to execute')